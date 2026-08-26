#!/usr/bin/env python3
"""Compare Amazon delivery manifests against a monthly bidding sheet."""

from __future__ import annotations

import argparse
import csv
import json
from collections import defaultdict
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook

from build_delivery_report_data import manifest_info, money, parse_decimal, parse_int, ratio, translate_gl


PRICE_BANDS = [
    (0, "<€5", Decimal("0"), Decimal("5")),
    (1, "€5-15", Decimal("5"), Decimal("15")),
    (2, "€15-25", Decimal("15"), Decimal("25")),
    (3, "€25-35", Decimal("25"), Decimal("35")),
    (4, "€35-45", Decimal("35"), Decimal("45")),
    (5, ">€45", Decimal("45"), None),
]

PRICE_BAND_ORDER = {label: order for order, label, _lower, _upper in PRICE_BANDS}


def price_band(unitcost: Decimal) -> tuple[int, str]:
    for order, label, lower, upper in PRICE_BANDS:
        if upper is None and unitcost >= lower:
            return order, label
        if unitcost >= lower and unitcost < upper:
            return order, label
    return PRICE_BANDS[0][0], PRICE_BANDS[0][1]


def condition_from_removal_reason(reason: str) -> str:
    return "Sellable" if reason == "Overstock" else "Unsellable"


def make_bucket() -> dict:
    return {
        "rows": 0,
        "units": 0,
        "sellable_units": 0,
        "unsellable_units": 0,
        "asins": set(),
        "unitcost_value": Decimal("0"),
        "actual_recovery": Decimal("0"),
        "expected_recovery": Decimal("0"),
        "cheap_units_lt5": 0,
        "cheap_value_lt5": Decimal("0"),
        "unitcost_weighted": Decimal("0"),
        "unitrecovery_weighted": Decimal("0"),
        "source_files": set(),
        "item_desc": "",
    }


def add_row(bucket: dict, row: dict, units: int, unitcost: Decimal, unitrecovery: Decimal, expected_recovery: Decimal) -> None:
    condition = row["condition"]
    bucket["rows"] += 1
    bucket["units"] += units
    if condition == "Sellable":
        bucket["sellable_units"] += units
    else:
        bucket["unsellable_units"] += units
    if row["asin"]:
        bucket["asins"].add(row["asin"])
    if not bucket["item_desc"] and row.get("item_desc"):
        bucket["item_desc"] = row["item_desc"]
    bucket["unitcost_value"] += unitcost * units
    bucket["actual_recovery"] += unitrecovery * units
    bucket["expected_recovery"] += expected_recovery
    bucket["unitcost_weighted"] += unitcost * units
    bucket["unitrecovery_weighted"] += unitrecovery * units
    if unitcost < Decimal("5"):
        bucket["cheap_units_lt5"] += units
        bucket["cheap_value_lt5"] += unitcost * units
    bucket["source_files"].add(row["source_file"])


def finalize_bucket(bucket: dict) -> dict:
    units = bucket["units"]
    unitcost_value = bucket["unitcost_value"]
    actual_recovery = bucket["actual_recovery"]
    expected_recovery = bucket["expected_recovery"]
    return {
        "rows": bucket["rows"],
        "units": units,
        "sellable_units": bucket["sellable_units"],
        "unsellable_units": bucket["unsellable_units"],
        "distinct_asins": len(bucket["asins"]),
        "avg_unitcost": money(bucket["unitcost_weighted"] / units) if units else 0,
        "avg_unitrecovery": money(bucket["unitrecovery_weighted"] / units) if units else 0,
        "unitcost_value": money(unitcost_value),
        "actual_recovery": money(actual_recovery),
        "expected_recovery": money(expected_recovery),
        "recovery_delta": money(actual_recovery - expected_recovery),
        "actual_recovery_rate": ratio(actual_recovery / unitcost_value) if unitcost_value else 0,
        "expected_recovery_rate": ratio(expected_recovery / unitcost_value) if unitcost_value else 0,
        "cheap_units_lt5": bucket["cheap_units_lt5"],
        "cheap_value_lt5": money(bucket["cheap_value_lt5"]),
        "cheap_unit_share": ratio(Decimal(bucket["cheap_units_lt5"]) / Decimal(units)) if units else 0,
        "cheap_value_share": ratio(bucket["cheap_value_lt5"] / unitcost_value) if unitcost_value else 0,
        "source_files": len(bucket["source_files"]),
        "item_desc": bucket["item_desc"],
    }


def parse_bidding_sheet(path: Path) -> tuple[dict[int, dict], dict[tuple[int, str], dict]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    capacities: dict[int, dict] = {}
    for row_index in range(5, ws.max_row + 1):
        gl_number = ws.cell(row_index, 2).value
        category = ws.cell(row_index, 3).value
        if not isinstance(gl_number, int) or not category:
            continue
        capacities[gl_number] = {
            "gl_number": gl_number,
            "bid_category": str(category),
            "subcategories_note": ws.cell(row_index, 4).value or "",
            "capacity_unsellable": int(ws.cell(row_index, 5).value or 0),
            "capacity_sellable": int(ws.cell(row_index, 6).value or 0),
        }

    rates: dict[tuple[int, str], dict] = {}
    for row_index in range(5, ws.max_row + 1):
        gl_number = ws.cell(row_index, 8).value
        category = ws.cell(row_index, 9).value
        band = ws.cell(row_index, 10).value
        if not isinstance(gl_number, int) or not category or not band:
            continue
        rates[(gl_number, str(band))] = {
            "gl_number": gl_number,
            "bid_category": str(category),
            "price_band": str(band),
            "rate_unsellable": Decimal(str(ws.cell(row_index, 11).value or 0)),
            "rate_sellable": Decimal(str(ws.cell(row_index, 12).value or 0)),
        }
    return capacities, rates


def status_for(row: dict) -> str:
    if not row["in_bid"]:
        return "Sin bidding"
    if row["units"] == 0:
        return "Sin recibido"
    if row["sellable_usage"] > 1 or row["unsellable_usage"] > 1:
        return "Exceso condicion"
    if row["total_usage"] > 1:
        return "Exceso total"
    return "Dentro capacidad"


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--delivery-json", required=True)
    parser.add_argument("--bidding-xlsx", required=True)
    parser.add_argument("--output-json", required=True)
    args = parser.parse_args()

    delivery_data = json.loads(Path(args.delivery_json).read_text(encoding="utf-8"))
    source_dir = Path(delivery_data["summary"]["source_dir"])
    capacities, rates = parse_bidding_sheet(Path(args.bidding_xlsx))

    by_gl = defaultdict(make_bucket)
    by_gl_band = defaultdict(make_bucket)
    by_gl_band_condition = defaultdict(make_bucket)
    by_asin = defaultdict(make_bucket)
    all_bucket = make_bucket()
    gl_meta: dict[int, dict] = {}
    missing_rate_units = 0

    files = sorted(source_dir.glob("Liq_FBA_WeeklyManifest_V3_*_ND72B.txt*"))
    for path in files:
        manifest_country, manifest_date = manifest_info(path)
        with path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
            reader = csv.DictReader(handle, delimiter="\t", quotechar='"', escapechar="\\")
            for raw in reader:
                gl_number = parse_int(raw.get("GL"))
                gl_desc = raw.get("GLDesc") or "(Sin GLDesc)"
                gl_es = translate_gl(gl_desc)
                asin = raw.get("Asin") or ""
                units = parse_int(raw.get("Units"))
                unitcost = parse_decimal(raw.get("UnitCost"))
                unitrecovery = parse_decimal(raw.get("UnitRecovery"))
                condition = condition_from_removal_reason(raw.get("RemovalReason") or "")
                band_order, band = price_band(unitcost)
                rate_info = rates.get((gl_number, band), {})
                bid_rate = rate_info.get(
                    "rate_sellable" if condition == "Sellable" else "rate_unsellable",
                    Decimal("0"),
                )
                if not rate_info:
                    missing_rate_units += units
                expected_recovery = unitcost * units * bid_rate
                row = {
                    "manifest_date": manifest_date,
                    "manifest_country": manifest_country,
                    "source_file": path.name,
                    "gl_number": gl_number,
                    "gl_desc": gl_desc,
                    "gl_es": gl_es,
                    "asin": asin,
                    "item_desc": raw.get("ItemDesc") or "",
                    "condition": condition,
                    "price_band": band,
                    "price_band_order": band_order,
                }
                gl_meta[gl_number] = {"gl_desc": gl_desc, "gl_es": gl_es}
                for bucket in (
                    all_bucket,
                    by_gl[gl_number],
                    by_gl_band[(gl_number, band_order, band)],
                    by_gl_band_condition[(gl_number, band_order, band, condition)],
                    by_asin[(asin, gl_number)],
                ):
                    add_row(bucket, row, units, unitcost, unitrecovery, expected_recovery)

    all_gl_numbers = sorted(set(capacities) | set(by_gl))
    gl_rows = []
    for gl_number in all_gl_numbers:
        bid = capacities.get(gl_number, {})
        meta = gl_meta.get(gl_number, {})
        actual = finalize_bucket(by_gl[gl_number])
        cap_sellable = bid.get("capacity_sellable", 0)
        cap_unsellable = bid.get("capacity_unsellable", 0)
        cap_total = cap_sellable + cap_unsellable
        row = {
            "gl_number": gl_number,
            "gl_desc": meta.get("gl_desc", ""),
            "gl_es": meta.get("gl_es", ""),
            "bid_category": bid.get("bid_category", ""),
            "in_bid": bool(bid),
            "capacity_sellable": cap_sellable,
            "capacity_unsellable": cap_unsellable,
            "capacity_total": cap_total,
            **actual,
            "sellable_usage": ratio(Decimal(actual["sellable_units"]) / Decimal(cap_sellable)) if cap_sellable else 0,
            "unsellable_usage": ratio(Decimal(actual["unsellable_units"]) / Decimal(cap_unsellable)) if cap_unsellable else 0,
            "total_usage": ratio(Decimal(actual["units"]) / Decimal(cap_total)) if cap_total else 0,
        }
        row["status"] = status_for(row)
        gl_rows.append(row)

    gl_rows.sort(key=lambda row: (row["bid_category"] or row["gl_es"], row["gl_number"]))

    rate_rows = []
    for (gl_number, band_order, band, condition), bucket in by_gl_band_condition.items():
        bid = capacities.get(gl_number, {})
        rate_info = rates.get((gl_number, band), {})
        actual = finalize_bucket(bucket)
        bid_rate = rate_info.get(
            "rate_sellable" if condition == "Sellable" else "rate_unsellable",
            Decimal("0"),
        )
        rate_rows.append(
            {
                "gl_number": gl_number,
                "bid_category": bid.get("bid_category", ""),
                "gl_desc": gl_meta.get(gl_number, {}).get("gl_desc", ""),
                "gl_es": gl_meta.get(gl_number, {}).get("gl_es", ""),
                "price_band_order": band_order,
                "price_band": band,
                "condition": condition,
                "bid_rate": float(bid_rate),
                **actual,
            }
        )
    rate_rows.sort(key=lambda row: (row["bid_category"], row["price_band_order"], row["condition"]))

    bidding_rate_rows = []
    for (gl_number, band), rate_info in rates.items():
        bid = capacities.get(gl_number, {})
        band_order = PRICE_BAND_ORDER.get(band, 99)
        actual = finalize_bucket(by_gl_band[(gl_number, band_order, band)])
        bidding_rate_rows.append(
            {
                "gl_number": gl_number,
                "bid_category": bid.get("bid_category", rate_info.get("bid_category", "")),
                "gl_desc": gl_meta.get(gl_number, {}).get("gl_desc", ""),
                "gl_es": gl_meta.get(gl_number, {}).get("gl_es", ""),
                "price_band_order": band_order,
                "price_band": band,
                "rate_unsellable": float(rate_info["rate_unsellable"]),
                "rate_sellable": float(rate_info["rate_sellable"]),
                **actual,
            }
        )
    bidding_rate_rows.sort(key=lambda row: (row["bid_category"], row["price_band_order"]))

    bid_strategy_rows = []
    for gl_number in sorted(capacities):
        bid = capacities.get(gl_number, {})
        actual = finalize_bucket(by_gl[gl_number])
        units_lt15 = 0
        value_lt15 = Decimal("0")
        units_ge35 = 0
        value_ge35 = Decimal("0")
        for band_order, band, _lower, _upper in PRICE_BANDS:
            bucket = by_gl_band[(gl_number, band_order, band)]
            if band_order <= 1:
                units_lt15 += bucket["units"]
                value_lt15 += bucket["unitcost_value"]
            if band_order >= 4:
                units_ge35 += bucket["units"]
                value_ge35 += bucket["unitcost_value"]

        rate_lt5 = rates.get((gl_number, "<€5"), {})
        rate_5_15 = rates.get((gl_number, "€5-15"), {})
        rate_35_45 = rates.get((gl_number, "€35-45"), {})
        rate_gt45 = rates.get((gl_number, ">€45"), {})
        cheap_share = ratio(Decimal(units_lt15) / Decimal(actual["units"])) if actual["units"] else 0
        high_share = ratio(Decimal(units_ge35) / Decimal(actual["units"])) if actual["units"] else 0
        value_high_share = ratio(value_ge35 / Decimal(str(actual["unitcost_value"]))) if actual["unitcost_value"] else 0
        rate_premium = float(rate_gt45.get("rate_sellable", Decimal("0")) - rate_lt5.get("rate_sellable", Decimal("0")))
        bid_strategy_rows.append(
            {
                "gl_number": gl_number,
                "bid_category": bid.get("bid_category", ""),
                "gl_desc": gl_meta.get(gl_number, {}).get("gl_desc", ""),
                "gl_es": gl_meta.get(gl_number, {}).get("gl_es", ""),
                "units": actual["units"],
                "unitcost_value": actual["unitcost_value"],
                "avg_unitcost": actual["avg_unitcost"],
                "rate_lt5_sellable": float(rate_lt5.get("rate_sellable", Decimal("0"))),
                "rate_5_15_sellable": float(rate_5_15.get("rate_sellable", Decimal("0"))),
                "rate_35_45_sellable": float(rate_35_45.get("rate_sellable", Decimal("0"))),
                "rate_gt45_sellable": float(rate_gt45.get("rate_sellable", Decimal("0"))),
                "rate_premium_gt45_vs_lt5": rate_premium,
                "units_lt15": units_lt15,
                "unit_share_lt15": cheap_share,
                "value_lt15": money(value_lt15),
                "units_ge35": units_ge35,
                "unit_share_ge35": high_share,
                "value_ge35": money(value_ge35),
                "value_share_ge35": value_high_share,
                "mismatch_score": round(rate_premium * cheap_share, 4),
            }
        )
    bid_strategy_rows.sort(key=lambda row: (-row["mismatch_score"], -row["units"], row["bid_category"]))

    asin_rows = []
    for (asin, gl_number), bucket in by_asin.items():
        if not asin:
            continue
        bid = capacities.get(gl_number, {})
        actual = finalize_bucket(bucket)
        asin_rows.append(
            {
                "asin": asin,
                "gl_number": gl_number,
                "bid_category": bid.get("bid_category", ""),
                "gl_desc": gl_meta.get(gl_number, {}).get("gl_desc", ""),
                "gl_es": gl_meta.get(gl_number, {}).get("gl_es", ""),
                **actual,
            }
        )
    asin_rows.sort(key=lambda row: (-row["units"], -row["unitcost_value"], row["asin"]))

    summary_actual = finalize_bucket(all_bucket)
    total_capacity_sellable = sum(row["capacity_sellable"] for row in capacities.values())
    total_capacity_unsellable = sum(row["capacity_unsellable"] for row in capacities.values())
    total_capacity = total_capacity_sellable + total_capacity_unsellable
    summary = {
        **summary_actual,
        "files": len(files),
        "bidding_file": str(Path(args.bidding_xlsx)),
        "source_dir": str(source_dir),
        "bid_categories": len(capacities),
        "matched_gl_categories": sum(1 for row in gl_rows if row["in_bid"] and row["units"] > 0),
        "missing_rate_units": missing_rate_units,
        "capacity_sellable": total_capacity_sellable,
        "capacity_unsellable": total_capacity_unsellable,
        "capacity_total": total_capacity,
        "sellable_usage": ratio(Decimal(summary_actual["sellable_units"]) / Decimal(total_capacity_sellable))
        if total_capacity_sellable
        else 0,
        "unsellable_usage": ratio(Decimal(summary_actual["unsellable_units"]) / Decimal(total_capacity_unsellable))
        if total_capacity_unsellable
        else 0,
        "total_usage": ratio(Decimal(summary_actual["units"]) / Decimal(total_capacity)) if total_capacity else 0,
        "comparison_note": "Overstock se clasifica como Sellable; Customer Damage, Defective, Vendor Damage y Carrier Damage como Unsellable.",
    }

    output = {
        "summary": summary,
        "gl_comparison": gl_rows,
        "bidding_rates": bidding_rate_rows,
        "bid_strategy": bid_strategy_rows,
        "rate_comparison": rate_rows,
        "top_asins": asin_rows[:500],
        "capacity_only": [row for row in gl_rows if row["units"] == 0],
        "not_in_bid": [row for row in gl_rows if not row["in_bid"] and row["units"] > 0],
    }

    output_path = Path(args.output_json)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Units: {summary['units']}")
    print(f"Sellable units: {summary['sellable_units']}")
    print(f"Unsellable units: {summary['unsellable_units']}")
    print(f"Total capacity usage: {summary['total_usage']:.1%}")
    print(f"Expected recovery: {summary['expected_recovery']}")
    print(f"Actual recovery: {summary['actual_recovery']}")
    print(f"Output: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
